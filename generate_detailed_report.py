#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
業績計画用レポートの詳細版を生成
指定された列のみを含む
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import random

def generate_detailed_report(filename="業績計画用レポート_詳細版.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "業績計画用レポート"

    # スタイル定義
    bold_font = Font(bold=True, size=9)
    normal_font = Font(size=9)
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    # 列の設定（A列から順に）
    # A列から順に並べるが、データを入れるのは指定列のみ
    # 実際の列位置：F=6, K=11, M=13, N=14, O=15, P=16, Q=17, R=18, S=19, T=20, U=21, AC=29, AD=30, AE=31, AF=32
    
    # 列幅設定（指定された列のみ）
    column_widths = {
        'F': 15,   # 営業案件タイプ
        'K': 12,   # 受注予定日
        'M': 12,   # フェーズ
        'N': 12,   # 
        'O': 12,   # 
        'P': 12,   # 
        'Q': 12,   # 
        'R': 12,   # 
        'S': 12,   # 
        'T': 12,   # 
        'U': 12,   # 
        'AC': 15,  # ステータス
        'AD': 15,  # 
        'AE': 12,  # 
        'AF': 12,  # 
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # タイトル行（2行目）
    ws['A2'] = "業績計画用レポート_今期（37下）_3部2G"
    ws['A2'].font = Font(size=14, bold=True)
    ws.merge_cells('A2:AF2')

    # メタデータ行（3行目）
    ws['A3'] = f"生成日時: {datetime.now().strftime('%Y/%m/%d %H:%M')}"
    ws['A3'].font = Font(size=9)

    # ヘッダー行（5行目）- スクショから推測される列名
    headers_row5 = {
        'A': '-',
        'B': '-',  # 折りたたまれている
        'C': '-',
        'D': '-',
        'E': '-',  # 折りたたまれている
        'F': '営業案件タイプ',
        'G': '-',
        'H': '-',
        'I': '-',
        'J': '-',
        'K': '受注予定日',
        'L': '-',
        'M': '営業担当者G',
        'N': '営業担当名',
        'O': '取引先：取引先名',
        'P': '取引先ID',
        'Q': '営業案件：案件名',
        'R': '契約確度',
        'S': '契約額37下（千円）',
        'T': '受注予定日',
        'U': 'チャネル',
        'V': '-',  # 折りたたまれている
        'W': '-',
        'X': '-',
        'Y': '-',
        'Z': '-',
        'AA': '-',
        'AB': '-',
        'AC': 'フェーズ',
        'AD': 'ステータス',
        'AE': '前年同期',
        'AF': '前年同期売上',
    }
    
    for col, header_text in headers_row5.items():
        cell = ws[f'{col}5']
        cell.value = header_text
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # サンプルデータ（20行程度）
    project_types = ['人材開発支援助', '人材開発支援助　人対面', '人材開発支援助　自社', '研修実施費助　助成額', '研修実施費助　助成額　受託・構内作業']
    companies = [
        '株式会社　ＮＨＫアイテス',
        '三菱ガス化学　株式会社',
        '五洋建設　株式会社',
        '日本電子計算　株式会社',
        '株式会社　そーう',
        '東洋電装　株式会社',
        '株式会社　ＭＩＲＩ',
        '株式会社　ＭＡＲＴＡＩＮＥ',
        'フリー　株式会社',
        '川崎重工業　株式会社',
    ]
    project_names = [
        'ロジカルシンキング',
        '新人フォロー施策',
        'マネージャー（早行）',
        'HRM-Q',
        'M-BT',
        'BCSP-M',
        'MOA',
        'オリジナルCRS',
        'ビジネス交渉',
        'JM BOOT CAMP',
    ]
    probabilities = ['Bヨミ', 'Cヨミ', 'ネタ']
    channels = ['①取引窓口×リ', '②取引窓口×新規']
    phases = ['未着商', '新規商談未開始', '着商', '提案中']
    statuses = ['進行中', '完了', '保留']

    # データ行生成（6行目から25行目まで）
    for row_idx in range(6, 26):
        # 指定された列のみデータを入れる
        row_data = {
            'A': '-',
            'B': '-',
            'C': '-',
            'D': '-',
            'E': '-',
            'F': random.choice(project_types),
            'G': '-',
            'H': '-',
            'I': '-',
            'J': '-',
            'K': f"2025/{random.randint(11, 12)}/{random.randint(1, 28):02d}",
            'L': '-',
            'M': '3部2 G',
            'N': '佐藤　温',
            'O': random.choice(companies),
            'P': f"a1VTL0000{random.randint(0, 9):d}{random.randint(100, 999):03d}",
            'Q': random.choice(project_names),
            'R': random.choice(probabilities),
            'S': random.choice([0, 700, 1200, 2600, 3000, 8000, 13000]),
            'T': f"2026/{random.randint(1, 3)}/{random.randint(1, 28):02d}",
            'U': random.choice(channels),
            'V': '-',
            'W': '-',
            'X': '-',
            'Y': '-',
            'Z': '-',
            'AA': '-',
            'AB': '-',
            'AC': random.choice(phases),
            'AD': random.choice(statuses),
            'AE': random.randint(2300, 2600),
            'AF': random.randint(2300, 2600),
        }
        
        for col, value in row_data.items():
            cell = ws[f'{col}{row_idx}']
            cell.value = value
            cell.font = normal_font
            cell.border = thin_border
            
            # アライメント設定
            if col in ['M', 'N', 'O', 'Q', 'U', 'AC', 'AD']:
                cell.alignment = left_align
            elif col in ['S', 'AE', 'AF']:
                cell.alignment = right_align
            else:
                cell.alignment = center_align
            
            # 赤色セル（フェーズが「未着商」「新規商談未開始」の場合）
            if col == 'AC' and value in ['未着商', '新規商談未開始']:
                cell.fill = red_fill

    # 行の高さ調整
    for row in range(1, 26):
        ws.row_dimensions[row].height = 18

    wb.save(filename)
    print(f"✅ 詳細版Excelファイルを生成しました: {filename}")
    print(f"   - 指定列: F, K, M～U, AC, AD, AE, AF")
    print(f"   - データ行数: 20行")
    print(f"   - その他の列は「-」で埋めています")

if __name__ == "__main__":
    generate_detailed_report()

