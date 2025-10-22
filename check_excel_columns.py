import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('業績計画用レポート_詳細版.xlsx')
ws = wb.active

print("=== Excelファイルの列構造確認 ===\n")

# F5セルを確認
print(f"F5セル: {ws['F5'].value}")
print(f"F列の列番号: 6 (A=1, B=2, ..., F=6)")
print(f"Python配列インデックス: F列=5 (0-based)\n")

# 5行目（ヘッダー行）の全列を表示
print("=== 5行目（ヘッダー行）の内容 ===")
for col in range(1, 30):
    col_letter = get_column_letter(col)
    cell_value = ws.cell(5, col).value
    if cell_value and str(cell_value).strip() and cell_value != '-':
        print(f"{col_letter}5 (配列[4][{col-1}]): {cell_value}")

print("\n=== 6行目（最初のデータ行）の内容 ===")
for col in range(1, 30):
    col_letter = get_column_letter(col)
    cell_value = ws.cell(6, col).value
    if cell_value and str(cell_value).strip() and cell_value != '-':
        print(f"{col_letter}6 (配列[5][{col-1}]): {cell_value}")

# M～R列の確認
print("\n=== M～R列（36上～38下）の確認 ===")
print(f"M5 (配列[4][12]): {ws['M5'].value}")
print(f"N5 (配列[4][13]): {ws['N5'].value}")
print(f"O5 (配列[4][14]): {ws['O5'].value}")
print(f"P5 (配列[4][15]): {ws['P5'].value}")
print(f"Q5 (配列[4][16]): {ws['Q5'].value}")
print(f"R5 (配列[4][17]): {ws['R5'].value}")

print("\n=== M～R列（6行目のデータ）===")
print(f"M6 (配列[5][12]): {ws['M6'].value}")
print(f"N6 (配列[5][13]): {ws['N6'].value}")
print(f"O6 (配列[5][14]): {ws['O6'].value}")
print(f"P6 (配列[5][15]): {ws['P6'].value}")
print(f"Q6 (配列[5][16]): {ws['Q6'].value}")
print(f"R6 (配列[5][17]): {ws['R6'].value}")

