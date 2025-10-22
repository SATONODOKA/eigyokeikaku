import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('業績計画用レポート_詳細版.xlsx')
ws = wb.active

print("=== ファイル情報 ===")
print(f"シート名: {ws.title}")
print(f"最大行: {ws.max_row}")
print(f"最大列: {ws.max_column}")

print("\n=== ヘッダー行（5行目）の列構成 ===")
for col in range(1, ws.max_column + 1):
    col_letter = get_column_letter(col)
    val = ws.cell(5, col).value
    if val and str(val).strip() and val != '-':
        print(f"{col_letter}列 (列{col}): {val}")

print("\n=== データ行サンプル（6～10行目） ===")
for row in range(6, 11):
    print(f"\n--- {row}行目 ---")
    row_data = []
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        val = ws.cell(row, col).value
        if val and str(val).strip() and val != '-':
            row_data.append(f"{col_letter}={val}")
    print(" | ".join(row_data) if row_data else "（空行）")

