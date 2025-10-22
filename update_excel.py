#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
既存のExcelファイルを編集して列配置を修正
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

def update_excel_layout():
    # 既存ファイルを読み込み
    filename = '業績計画用レポート_ダミー.xlsx'
    wb = load_workbook(filename)
    ws = wb.active
    
    # 既存データをクリア（行15以降）
    ws.delete_rows(15, ws.max_row - 14)
    
    # 列幅の設定（A列は空白、B列からデータ開始）
    ws.column_dimensions['A'].width = 2   # 空白列
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 28
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 18
    ws.column_dimensions['K'].width = 20
    
    # ヘッダー行（行15） - B列から開始
    headers = [
        '',  # A列は空白
        '営業担当者G',
        '',  # C列は結合するので空
        '営業担当名',
        '取引先：取引先名',
        '営業案件：案件名',
        '契約確度',
        '契約額37下（千円）',
        '受注予定日',
        'チャネル',
        '営業案件：ID'
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=15, column=col, value=header)
        if col > 1:  # A列以外
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # B列とC列を結合（ヘッダー行）
    ws.merge_cells('B15:C15')
    
    # データ行（行16以降） - スクショに合わせてデータを配置
    data_rows = [
        ['', '3部2 G', '', '佐藤　温', 'フリー　株式会社', 'JM BOOT CAMP（3月Bヨミ', 'Bヨミ', '', '2025/12/31', '①取引窓口×リ', 'a1VTL00000DtpC5'],
        ['', '3部2 G', '', '佐藤　温', '株式会社　そーう　り）', '3年目研修', 'Cヨミ', 'JPY700', '2025/12/26', '①取引窓口×リ', 'a1VF900001Y1i7j'],
        ['', '3部2 G', '', '佐藤　温', '三菱ガス化学　株式会社', '新人フォロー施策', 'Cヨミ', '', '2025/11/28', '②取引窓口×新規', 'a1VF900001Z6Rbw'],
        ['', '3部2 G', '', '佐藤　温', '日本電子計算　株式会社', 'HRM-Q', 'ネタ', '', '2025/11/28', '①取引窓口×リ', 'a1VTL00000Gcsat'],
        ['', '3部2 G', '', '佐藤　温', '日本電子計算　株式会社', 'M-BT', 'ネタ', '', '2025/11/28', '①取引窓口×リ', 'a1VTL00000Gcsau'],
        ['', '3部2 G', '', '佐藤　温', 'グ環境サービス　経営幹部向けCO-WS', '', 'ネタ', 'JPY500', '2026/03/31', '②取引窓口×新規', 'a1VTL00000Gcu5z'],
        ['', '3部2 G', '', '佐藤　温', '株式会社　ＭＡＲＴＡＩＮＥ', '', 'ネタ', '', '2026/01/30', '①取引窓口×リ', 'a1VTL00000GcuzF'],
        ['', '3部2 G', '', '佐藤　温', '株式会社　ＭＩＲＩＳＴＡＲ', '', 'ネタ', '', '2026/01/30', '①取引窓口×リ', 'a1VTL00000Gcv45'],
        ['', '3部2 G', '', '佐藤　温', '株式会社　ＭＩＲＩ', 'オリジナルCRS', 'ネタ', '', '2026/01/30', '①取引窓口×リ', 'a1VTL00000GcvIb'],
        ['', '3部2 G', '', '佐藤　温', '株式会社　ＭＩＲＩ', 'ビジネス交渉', 'ネタ', '', '2026/01/30', '①取引窓口×リ', 'a1VTL00000GcvNR'],
        ['', '3部2 G', '', '佐藤　温', '五洋建設　株式会社', 'BCSP-M', 'ネタ', '', '2025/12/31', '①取引窓口×リ', 'a1VTL00000GcwMl'],
        ['', '3部2 G', '', '佐藤　温', '株式会社　そーう', 'マネジメント基礎研修', 'ネタ', '', '2026/03/31', '①取引窓口×リ', 'a1VTL00000GcxnR'],
        ['', '3部2 G', '', '佐藤　温', '東洋電装　株式会社', 'MOA', 'ネタ', '', '2026/02/27', '①取引窓口×リ', 'a1VTL00000GcxvV'],
        ['', '3部2 G', '', '佐藤　温', '東洋電装　株式会社', 'MOA-FBS', 'ネタ', '', '2026/02/27', '①取引窓口×リ', 'a1VTL00000Gcy3Z'],
        ['', '3部2 G', '', '佐藤　温', '株式会社　新日本マネジメント', '研修', 'ネタ', '', '2026/02/27', '①取引窓口×リ', 'a1VTL00000GczE9'],
        ['', '3部2 G', '', '佐藤　温', '日本電子計算　株式会社', 'カスタムアセス', 'ネタ', '', '2026/01/30', '①取引窓口×リ', 'a1VTL00000GczUl'],
        ['', '3部2 G', '', '佐藤　温', '三菱ガス化学　株式会社', 'M-BT', 'ネタ', '', '2025/12/31', '①取引窓口×リ', 'a1VTL00000GczkP'],
        ['', '3部2 G', '', '佐藤　温', '三菱ガス化学　株式会社', 'F-BT', 'ネタ', '', '2025/11/28', '①取引窓口×リ', 'a1VTL00000GczxJ'],
        ['', '3部2 G', '', '佐藤　温', '三菱ガス化学　株式会社', 'MINE-SM', 'ネタ', '', '2025/12/31', '①取引窓口×リ', 'a1VTL00000Gd0Oj']
    ]
    
    for row_idx, row_data in enumerate(data_rows, start=16):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='left', vertical='center')
        # B列とC列を結合（データ行）
        ws.merge_cells(f'B{row_idx}:C{row_idx}')
    
    # フッター行（最後の行）
    last_row = 16 + len(data_rows)
    ws[f'B{last_row}'] = f'業績計画用レポート_今期（37下）_3部2G'
    ws[f'B{last_row}'].font = Font(size=9)
    
    # ファイル保存
    wb.save(filename)
    print(f'✅ Excelファイルを更新しました: {filename}')
    print(f'   - A列: 空白')
    print(f'   - B-C列: 営業担当者G')
    print(f'   - D列: 営業担当名')
    print(f'   - E列: 取引先：取引先名')
    print(f'   - F列: 営業案件：案件名')
    print(f'   - G列: 契約確度（Bヨミ、Cヨミ、ネタ）')
    print(f'   - H列: 契約額37下（千円）')
    print(f'   - I列: 受注予定日')
    print(f'   - J列: チャネル')
    print(f'   - K列: 営業案件：ID')
    print(f'   - データ行数: {len(data_rows)}行')
    return filename

if __name__ == '__main__':
    update_excel_layout()

