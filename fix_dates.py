import openpyxl
from openpyxl.styles import Alignment

# Excelファイルを開く
wb = openpyxl.load_workbook('業績計画用レポート_詳細版.xlsx')
ws = wb.active

# 期ごとの代表的な受注年月（YYMM形式）
period_dates = {
    'M': 2404,  # 36上 = 2024年4月（数値型）
    'N': 2410,  # 36下 = 2024年10月
    'O': 2504,  # 37上 = 2025年4月
    'P': 2510,  # 37下 = 2025年10月
    'Q': 2604,  # 38上 = 2026年4月
    'R': 2610   # 38下 = 2026年10月
}

# 列の対応関係
column_mapping = {
    'M': 'AA',  # M列 → AA列
    'N': 'AB',  # N列 → AB列
    'O': 'AC',  # O列 → AC列
    'P': 'AD',  # P列 → AD列
    'Q': 'AE',  # Q列 → AE列
    'R': 'AF'   # R列 → AF列
}

# 6行目から最終行まで処理
count_set = 0
count_clear = 0
for row in range(6, ws.max_row + 1):
    for amount_col, date_col in column_mapping.items():
        # 金額セル（M～R列）の値を取得
        amount_cell = ws[f'{amount_col}{row}']
        amount_value = amount_cell.value
        
        # 日付セル
        date_cell = ws[f'{date_col}{row}']
        
        # 金額が1以上の場合、対応する受注年月を設定（数値型、右寄せ）
        if amount_value and isinstance(amount_value, (int, float)) and amount_value >= 1:
            date_cell.value = period_dates[amount_col]
            date_cell.number_format = '0'  # 数値として表示
            date_cell.alignment = Alignment(horizontal='right')
            count_set += 1
        else:
            # 金額が0または空の場合、日付セルを空欄に
            date_cell.value = None
            count_clear += 1

# M～R列も右寄せに設定
for row in range(6, ws.max_row + 1):
    for col in ['M', 'N', 'O', 'P', 'Q', 'R']:
        cell = ws[f'{col}{row}']
        if cell.value is not None:
            cell.alignment = Alignment(horizontal='right')

print(f'✅ {count_set}個のセルに受注年月を設定しました')
print(f'✅ {count_clear}個のセルを空欄にしました')
print('✅ 数値を右寄せに設定しました')

# 保存
wb.save('業績計画用レポート_詳細版.xlsx')
print('✅ ファイルを保存しました')

