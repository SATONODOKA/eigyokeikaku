#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
業績計画用レポートのダミーExcelファイル生成スクリプト
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

def create_dummy_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "業績計画用レポート_今期（37下）_3部2G"
    
    # 列幅の設定
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    
    # タイトル行（行2）
    ws['A2'] = '業績計画用レポート_今期（37下）_3部2G'
    ws['A2'].font = Font(size=14, bold=True)
    
    # メタデータ行（行3）
    ws['A3'] = '2025-10-22 15:51:07 日本標準時/JSTの時点 • 生成者: 佐藤 温 • 並び替え基準: 契約確度、昇順'
    ws['A3'].font = Font(size=9)
    
    # 検索条件セクション
    ws['A6'] = '検索条件'
    ws['A6'].font = Font(bold=True)
    
    ws['A7'] = '表示: すべての営業案件'
    ws['A8'] = '日付項目: 受注予定日 次の文字列と一致する カスタム (2025/10/01 〜 2026/03/31)'
    ws['A9'] = '営業案件: レコードタイプ 次の文字列と一致する 案件'
    ws['A10'] = '契約確度 次の文字列を含まない 実績'
    ws['A11'] = 'フェーズ 次の文字列と一致しない 受注,没,実績,目標,計画'
    ws['A12'] = '営業担当者G 次の文字列と一致する 3部2 G'
    ws['A13'] = '営業担当者名 次の文字列と一致する 佐藤　温'
    
    # ヘッダー行（行15）
    headers = [
        '営業担当者G',
        '営業担当名',
        '取引先',
        '営業案件: 案件名',
        'チャネル',
        '契約額37下（千円）',
        '受注予定日',
        '営業案件: ID'
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=15, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # データ行（行16以降）
    data_rows = [
        ['3部2 G', '佐藤　温', 'フリー　株式会社', 'JM BOOT CAMP（3月Bヨミ', '①取引窓口×リ', '', '2025/12/31', 'a1VTL00000DtpC5'],
        ['3部2 G', '佐藤　温', '株式会社　そーう　り）', '3年目研修', 'Cヨミ', 'JPY700', '2025/12/26', 'a1VF900001Y1i7j'],
        ['3部2 G', '佐藤　温', '三菱ガス化学　株式会社', '新人フォロー施策', 'Cヨミ', '', '2025/11/28', 'a1VF900001Z6Rbw'],
        ['3部2 G', '佐藤　温', '日本電子計算　株式会社', 'HRM-Q', 'ネタ', '', '2025/11/28', 'a1VTL00000Gcsat'],
        ['3部2 G', '佐藤　温', '日本電子計算　株式会社', 'M-BT', 'ネタ', '', '2025/11/28', 'a1VTL00000Gcsau'],
        ['3部2 G', '佐藤　温', 'グ環境サービス', '経営幹部向けCO-WS', 'ネタ', 'JPY500', '2026/03/31', 'a1VTL00000Gcu5z'],
        ['3部2 G', '佐藤　温', '株式会社　ＭＡＲＴＡＩＮＥ', '', 'ネタ', '', '2026/01/30', 'a1VTL00000GcuzF'],
        ['3部2 G', '佐藤　温', '株式会社　ＭＩＲＩＳＴＡＲ', '', 'ネタ', '', '2026/01/30', 'a1VTL00000Gcv45'],
        ['3部2 G', '佐藤　温', '株式会社　ＭＩＲＩ', 'オリジナルCRS', 'ネタ', '', '2026/01/30', 'a1VTL00000GcvIb'],
        ['3部2 G', '佐藤　温', '株式会社　ＭＩＲＩ', 'ビジネス交渉', 'ネタ', '', '2026/01/30', 'a1VTL00000GcvNR'],
        ['3部2 G', '佐藤　温', '五洋建設　株式会社', 'BCSP-M', 'ネタ', '', '2025/12/31', 'a1VTL00000GcwMl'],
        ['3部2 G', '佐藤　温', '株式会社　そーう', 'マネジメント基礎研修', 'ネタ', '', '2026/03/31', 'a1VTL00000GcxnR'],
        ['3部2 G', '佐藤　温', '東洋電装　株式会社', 'MOA', 'ネタ', '', '2026/02/27', 'a1VTL00000GcxvV'],
        ['3部2 G', '佐藤　温', '東洋電装　株式会社', 'MOA-FBS', 'ネタ', '', '2026/02/27', 'a1VTL00000Gcy3Z'],
        ['3部2 G', '佐藤　温', '株式会社　新日本マネジメント', '研修', 'ネタ', '', '2026/02/27', 'a1VTL00000GczE9'],
        ['3部2 G', '佐藤　温', '日本電子計算　株式会社', 'カスタムアセス', 'ネタ', '', '2026/01/30', 'a1VTL00000GczUl'],
        ['3部2 G', '佐藤　温', '三菱ガス化学　株式会社', 'M-BT', 'ネタ', '', '2025/12/31', 'a1VTL00000GczkP'],
        ['3部2 G', '佐藤　温', '三菱ガス化学　株式会社', 'F-BT', 'ネタ', '', '2025/11/28', 'a1VTL00000GczxJ'],
        ['3部2 G', '佐藤　温', '三菱ガス化学　株式会社', 'MINE-SM', 'ネタ', '', '2025/12/31', 'a1VTL00000Gd0Oj']
    ]
    
    for row_idx, row_data in enumerate(data_rows, start=16):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # フッター行（最後の行）
    last_row = 16 + len(data_rows)
    ws[f'A{last_row}'] = f'業績計画用レポート_今期（37下）_3部2G'
    ws[f'A{last_row}'].font = Font(size=9)
    
    # ファイル保存
    filename = '業績計画用レポート_ダミー.xlsx'
    wb.save(filename)
    print(f'✅ Excelファイルを生成しました: {filename}')
    print(f'   - データ行数: {len(data_rows)}行')
    print(f'   - 列数: {len(headers)}列')
    return filename

if __name__ == '__main__':
    create_dummy_excel()

