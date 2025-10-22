import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('業績計画用レポート_詳細版.xlsx')
ws = wb.active

print("=== 最新ファイル確認 ===")
print(f"最大行: {ws.max_row}")

print("\n=== ヘッダー行（5行目） ===")
for col in [3, 4, 6, 7, 13, 14, 15, 16, 17, 18, 22, 27, 28, 29, 30]:
    col_letter = get_column_letter(col)
    val = ws.cell(5, col).value
    print(f"{col_letter}列: {val}")

print("\n=== サンプルデータ（6～8行目） ===")
for row in range(6, 9):
    print(f"\n{row}行目:")
    print(f"  F列（会社名）: {ws.cell(row, 6).value}")
    print(f"  M列（36上）: {ws.cell(row, 13).value}")
    print(f"  N列（36下）: {ws.cell(row, 14).value}")
    print(f"  O列（37上）: {ws.cell(row, 15).value}")
    print(f"  P列（37下）: {ws.cell(row, 16).value}")
    print(f"  Q列（38上）: {ws.cell(row, 17).value}")
    print(f"  R列（38下）: {ws.cell(row, 18).value}")

